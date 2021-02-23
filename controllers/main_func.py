from PIL import Image
from io import BytesIO
import requests
import openpyxl


def get_img(url, size=(100, 100)):
    r = requests.get(url, stream=True)
    if not r.ok:
        r.raise_for_error()
    r.raw.decode_content = True
    img = Image.open(r.raw)
    if size:
        img = img.resize(size)
    temp = BytesIO()
    img.save(temp, format="png")
    temp.seek(0)
    return Image.open(temp)


def insert_row(ws, info_good, size=(200, 200)):
    img = openpyxl.drawing.image.Image(get_img(info_good['photo'], size=size))
    row_num = ws.max_row + 1
    cell_addr = f"A{row_num}"
    img.anchor = cell_addr
    ws.add_image(img)
    ws[f"B{row_num}"] = info_good['title']
    ws[f"C{row_num}"] = info_good['brand']
    ws[f"D{row_num}"] = info_good['price']
    ws[f"E{row_num}"] = info_good['rating']
    ws[f"F{row_num}"] = info_good['reviews']
    ws[f"G{row_num}"] = info_good['quantity']
    ws[f"H{row_num}"] = info_good['url']
    ws.row_dimensions[row_num].height = int(size[1] * .8)
    ws.column_dimensions["A"].width = int(size[0] * .2)